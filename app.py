import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    
    # Check if the cell's value can be converted to a float
    try:
        cell_value = float(cell.value)
    except (ValueError, TypeError):
        # Handle non-numeric values or errors by printing and skipping
        print(f"Skipping row {row}: Value '{cell.value}' is not numeric.")
        continue
    
    corrected_price = (float(cell.value) * 0.9)
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, 
          min_row=2, 
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart= BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')






# import openpyxl as xl
# wb = xl.load_workbook('transactions.xlsx')
# sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1,1)

# for row in range(2,sheet.max_row+1):
#     cell = sheet.cell(row,3)

#     #   try:
#     #      cell_value = float(cell.value)
#     #  except (ValueError, TypeError):
#     #      cell_value = 0.0 

#     corrected_price = (float(cell.value) * 0.9)
#     corrected_price_cell = sheet.cell(row,4)
#     corrected_price_cell.value = corrected_price

# wb.save('transactions.xlsx')




